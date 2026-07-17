from __future__ import annotations

from openpyxl import Workbook
from pathlib import Path
from utils.synthetic_data.core.document_rendering.context import (
    GeneratedDocument,
    RenderContext,
)


def render_xlsx(ctx: RenderContext, *, output_dir: Path) -> GeneratedDocument:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{ctx.recipe.recipe_id}.xlsx"
    marker_en, marker_local = ctx.marker_lines

    workbook = Workbook()
    workbook.properties.title = ctx.country.name
    workbook.properties.description = marker_en

    marking_sheet = workbook.active
    marking_sheet.title = "notice"
    marking_sheet.append([marker_en])
    marking_sheet.append([marker_local])
    marking_sheet.sheet_view.rightToLeft = ctx.is_rtl

    metrics_sheet = workbook.create_sheet(title="metrics")
    metrics_sheet.sheet_view.rightToLeft = ctx.is_rtl
    metrics_sheet.append(["metric", "value"])
    for metric, value in sorted(ctx.country.current_metrics.items()):
        metrics_sheet.append([metric, value])

    text_sheet = workbook.create_sheet(title="blocks")
    text_sheet.sheet_view.rightToLeft = ctx.is_rtl
    text_sheet.append(["block_id", "text"])
    for block in ctx.recipe.blocks:
        text_sheet.append([block.block_id, block.text])

    workbook.save(str(path))
    return GeneratedDocument(
        path=path,
        file_format="xlsx",
        mode="n/a",
        locale=ctx.recipe.locale,
        country_id=ctx.country.country_id,
        recipe_id=ctx.recipe.recipe_id,
        related_artifact_ids=(ctx.country.country_id,),
        size_bytes=path.stat().st_size,
    )
