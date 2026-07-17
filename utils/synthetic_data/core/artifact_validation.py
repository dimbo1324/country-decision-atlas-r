from __future__ import annotations

import json
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from utils.synthetic_data.core.document_rendering.context import (
    GeneratedDocument,
)
from utils.synthetic_data.core.world_models import FICTIONAL_NOTICE


def _validate_json(document: GeneratedDocument) -> list[str]:
    payload = json.loads(document.path.read_text(encoding="utf-8"))
    errors = []
    if not payload.get("blocks"):
        errors.append(f"{document.path}: json has no blocks")
    return errors


def _validate_txt(document: GeneratedDocument) -> list[str]:
    text = document.path.read_text(encoding="utf-8")
    if FICTIONAL_NOTICE not in text:
        return [f"{document.path}: txt is missing the fictional notice"]
    return []


def _validate_xlsx(document: GeneratedDocument) -> list[str]:
    workbook = load_workbook(document.path)
    expected = {"notice", "metrics", "blocks"}
    if set(workbook.sheetnames) != expected:
        return [
            f"{document.path}: xlsx sheets {workbook.sheetnames} != {expected}"
        ]
    return []


def _validate_docx(document: GeneratedDocument) -> list[str]:
    docx_document = Document(str(document.path))
    if document.mode == "copyable" and not any(
        paragraph.text.strip() for paragraph in docx_document.paragraphs
    ):
        return [f"{document.path}: copyable docx has no extractable text"]
    return []


def _validate_pdf(document: GeneratedDocument) -> list[str]:
    reader = PdfReader(document.path)
    errors = []
    if not reader.pages:
        errors.append(f"{document.path}: pdf has no pages")
        return errors
    extracted = " ".join(page.extract_text() for page in reader.pages).strip()
    if document.mode == "copyable" and not extracted:
        errors.append(f"{document.path}: copyable pdf has no extractable text")
    if document.mode == "non_copyable" and extracted:
        errors.append(
            f"{document.path}: non_copyable pdf unexpectedly has "
            "extractable text"
        )
    return errors


_VALIDATORS = {
    "json": _validate_json,
    "txt": _validate_txt,
    "xlsx": _validate_xlsx,
    "docx": _validate_docx,
    "pdf": _validate_pdf,
}


def validate_generated_documents(
    documents: list[GeneratedDocument],
) -> list[str]:
    """Structural post-generation checks (spec section 14.4): every
    artifact opens with its normal library and has the properties its
    format/mode claims (extractable text for copyable, none for
    non_copyable, valid JSON, expected XLSX sheets)."""
    errors: list[str] = []
    for document in documents:
        validator = _VALIDATORS.get(document.file_format)
        if validator is None:
            continue
        try:
            errors.extend(validator(document))
        except Exception as error:
            errors.append(f"{document.path}: failed to open ({error})")
    return errors
