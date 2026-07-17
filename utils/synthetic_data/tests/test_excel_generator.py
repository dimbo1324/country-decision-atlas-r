from __future__ import annotations

from openpyxl import load_workbook
from pathlib import Path
from utils.synthetic_data.core.random_content import RandomContentFactory
from utils.synthetic_data.generators.excel_generator import ExcelGenerator


def test_excel_generator_writes_two_sheets_with_rows(
    tmp_path: Path, content: RandomContentFactory
) -> None:
    generator = ExcelGenerator()
    artifact = generator.generate(output_dir=tmp_path, content=content)

    workbook = load_workbook(artifact.path)
    assert len(workbook.sheetnames) == 2
    for sheet in workbook.worksheets:
        assert sheet.max_row > 1
        assert sheet.max_column == 5
