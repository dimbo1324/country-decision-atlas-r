from __future__ import annotations

import random
from openpyxl import load_workbook
from pathlib import Path
from scripts.synthetic_data.generators.excel_generator import ExcelGenerator


def test_excel_generator_writes_two_sheets_with_rows(tmp_path: Path) -> None:
    generator = ExcelGenerator()
    artifact = generator.generate(output_dir=tmp_path, rng=random.Random(3))

    workbook = load_workbook(artifact.path)
    assert len(workbook.sheetnames) == 2
    for sheet in workbook.worksheets:
        assert sheet.max_row > 1
        assert sheet.max_column == 5
